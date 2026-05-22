from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Places By Function"

headers = ["name", "function", "tech era"]
ws.append(headers)

base_entries = [
    ("Teepee Hut", "where people sleep and eat food sometimes", "tribal"),
    ("Spaceship", "primarily used to travel space", "spacer"),
    ("Press Skyscraper", "used for newspaper manufacturing, news stations, and media", "industrial"),
    ("Grocery Store", "used to store food stuffs and sell it", "industrial"),
    ("Queen Zone", "holds the queen. she gets fed people or something lol", "alien"),
]

for entry in base_entries:
    ws.append(entry)

eras = {
    "wild": ["Cave", "Fire Pit", "Hunter Camp", "Stone Cache", "River Shelter"],
    "tribal": ["Longhouse", "Drum Circle", "Sweat Lodge", "Clan Hut", "Watch Post"],
    "medieval": ["Castle", "Tavern", "Blacksmith", "Marketplace", "Barracks"],
    "industrial": ["Factory", "Hospital", "Train Station", "Warehouse", "Apartment Complex"],
    "ultratech": ["Nanoforge", "Teleport Hub", "AI Core", "Drone Hangar", "Cryostasis Vault"],
    "spacer": ["Orbital Station", "Starport", "Asteroid Mine", "Hydroponics Deck", "Cargo Freighter"],
    "alien": ["Egg Chamber", "Hive Tunnel", "Biomass Pit", "Flesh Spire", "Parasite Nursery"],
}

functions = {
    "wild": "used for survival, shelter, or food gathering",
    "tribal": "used by tribal communities for living, rituals, or defense",
    "medieval": "used by medieval societies for trade, defense, or daily life",
    "industrial": "used for manufacturing, transportation, or urban services",
    "ultratech": "used for advanced science, automation, or futuristic living",
    "spacer": "used for space travel, colonization, or orbital industry",
    "alien": "organic or unnatural structure used by alien lifeforms",
}

for era, names in eras.items():
    for i in range(1, 31):
        for name in names:
            ws.append((f"{name} {i}", functions[era], era))

for cell in ws[1]:
    cell.font = Font(bold=True)

for col in range(1, 4):
    max_len = 0
    col_letter = get_column_letter(col)
    for cell in ws[col_letter]:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

path = "places_by_function.xlsx"
wb.save(path)

print("Spreadsheet created successfully.")
print(f"Saved to: {path}")
