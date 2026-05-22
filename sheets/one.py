from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

entries = [
    ("House", "where people sleep, store belongings, and prepare food", "universal"),
]

# Allowed eras from user
eras = ["wild", "tribal", "medieval", "industrial", "ultratech", "spacer", "alien"]

data = [
    # Wild
    ("Cave Shelter", "used for protection from weather and predators", "wild"),
    ("Hunter Camp", "temporary resting place for hunting groups", "wild"),
    ("Berry Grove", "area where people gather edible plants", "wild"),
    ("Fishing Shore", "place used for catching fish", "wild"),
    ("Animal Trap Pit", "used to capture animals for food", "wild"),
    ("Fire Circle", "used for warmth, cooking, and social gatherings", "wild"),
    ("Stone Cache", "used to store gathered food and tools", "wild"),
    ("Mud Shelter", "simple shelter made from earth and branches", "wild"),
    ("River Crossing", "used to safely move across water", "wild"),
    ("Lookout Hill", "used to watch for danger or prey", "wild"),

    # Tribal
    ("Teepee Hut", "where people sleep and eat food sometimes", "tribal"),
    ("Longhouse", "shared living place for large families or clans", "tribal"),
    ("Totem Grounds", "used for ceremonies and spiritual gatherings", "tribal"),
    ("Smoke Hut", "used to preserve meat through smoking", "tribal"),
    ("Chief's Tent", "home and meeting place for tribal leaders", "tribal"),
    ("Tribal Forge", "used to craft primitive metal tools", "tribal"),
    ("Sweat Lodge", "used for cleansing rituals and ceremonies", "tribal"),
    ("Watch Post", "used to monitor nearby territory", "tribal"),
    ("War Drum Circle", "used to coordinate raids and gatherings", "tribal"),
    ("Trading Camp", "used to exchange goods between tribes", "tribal"),

    # Medieval
    ("Castle", "fortified home for nobles and soldiers", "medieval"),
    ("Blacksmith", "used to forge weapons, armor, and tools", "medieval"),
    ("Village Well", "used to collect drinking water", "medieval"),
    ("Tavern", "used for food, drink, and social gatherings", "medieval"),
    ("Stable", "used to house horses and transport animals", "medieval"),
    ("Watchtower", "used to defend territory and spot enemies", "medieval"),
    ("Monastery", "used for religious study and worship", "medieval"),
    ("Marketplace", "used for buying and selling goods", "medieval"),
    ("Siege Workshop", "used to build siege weapons", "medieval"),
    ("Dungeon", "used to imprison criminals or enemies", "medieval"),
    ("Windmill", "used to grind grain into flour", "medieval"),
    ("Apothecary", "used to prepare medicines and remedies", "medieval"),
    ("Town Hall", "used for governing a settlement", "medieval"),
    ("Barracks", "used to house soldiers", "medieval"),
    ("Cathedral", "large religious structure for worship", "medieval"),

    # Industrial
    ("Press Skyscraper", "used for newspaper manufacturing, news stations, and media", "industrial"),
    ("Grocery Store", "used to store food stuffs and sell it", "industrial"),
    ("Factory", "used for mass production of goods", "industrial"),
    ("Train Station", "used for rail transportation", "industrial"),
    ("Power Plant", "used to generate electricity", "industrial"),
    ("Hospital", "used to treat injured and sick people", "industrial"),
    ("Police Station", "used to enforce laws and hold suspects", "industrial"),
    ("Warehouse", "used to store industrial goods and supplies", "industrial"),
    ("Apartment Complex", "used to house many families in one structure", "industrial"),
    ("Office Tower", "used for administration and business work", "industrial"),
    ("School", "used for education and teaching", "industrial"),
    ("Shopping Mall", "used for large scale retail trade", "industrial"),
    ("Airport", "used for air travel and cargo", "industrial"),
    ("Data Center", "used to store and process computer information", "industrial"),
    ("Oil Refinery", "used to process crude oil into fuels", "industrial"),
    ("Cinema", "used for showing movies and entertainment", "industrial"),
    ("Subway Station", "used for underground public transportation", "industrial"),
    ("Bank", "used for storing money and handling financial services", "industrial"),
    ("Research Laboratory", "used for scientific experiments and testing", "industrial"),
    ("Prison", "used to confine criminals", "industrial"),

    # Ultratech
    ("Nanoforge", "used to manufacture advanced materials with nanotechnology", "ultratech"),
    ("Clone Vat Facility", "used to grow cloned organisms", "ultratech"),
    ("Arcology", "massive self contained city structure", "ultratech"),
    ("Cybernetic Clinic", "used to install artificial body parts", "ultratech"),
    ("Fusion Reactor", "used to generate high output clean energy", "ultratech"),
    ("Drone Hangar", "used to store and maintain autonomous drones", "ultratech"),
    ("Teleport Hub", "used for instant transportation between locations", "ultratech"),
    ("AI Core Chamber", "used to house advanced artificial intelligences", "ultratech"),
    ("Cryostasis Vault", "used to preserve people in suspended animation", "ultratech"),
    ("Hologram Theater", "used for immersive holographic entertainment", "ultratech"),
    ("Bio Lab", "used to engineer advanced organisms and medicine", "ultratech"),
    ("Orbital Elevator", "used to transport cargo into orbit", "ultratech"),
    ("Shield Generator", "used to create defensive energy barriers", "ultratech"),
    ("Quantum Server Farm", "used for high speed computation and simulations", "ultratech"),
    ("Synthetic Food Plant", "used to mass produce artificial food", "ultratech"),

    # Spacer
    ("Spaceship", "primarily used to travel space", "spacer"),
    ("Orbital Station", "used for living and working in orbit", "spacer"),
    ("Asteroid Mine", "used to extract minerals from asteroids", "spacer"),
    ("Cryo Bay", "used to store sleeping travelers during long journeys", "spacer"),
    ("Docking Ring", "used for connecting spacecraft together", "spacer"),
    ("Terraforming Array", "used to alter planetary environments", "spacer"),
    ("Starport", "used for launching and receiving spacecraft", "spacer"),
    ("Hydroponics Deck", "used to grow food in space habitats", "spacer"),
    ("Deep Space Observatory", "used to study distant space objects", "spacer"),
    ("FTL Relay", "used for faster than light communication", "spacer"),
    ("Vacuum Foundry", "used to manufacture materials in zero gravity", "spacer"),
    ("Generation Ship", "massive ship designed for multi generation travel", "spacer"),
    ("Moon Colony Dome", "pressurized settlement for living on moons", "spacer"),
    ("Cargo Freighter", "used to transport goods through space", "spacer"),
    ("Planetary Defense Cannon", "used to defend planets from space threats", "spacer"),

    # Alien
    ("Queen Zone", "holds the queen. she gets fed people or something lol", "alien"),
    ("Biomass Pit", "used to dissolve organic matter into nutrients", "alien"),
    ("Egg Chamber", "used to grow and store alien eggs", "alien"),
    ("Hive Tunnel", "used for movement inside a living hive", "alien"),
    ("Flesh Spire", "living tower used for communication and control", "alien"),
    ("Parasite Nursery", "used to raise parasitic organisms", "alien"),
    ("Bone Garden", "used to display remains of prey", "alien"),
    ("Resin Nest", "used to imprison captured creatures", "alien"),
    ("Mind Pool", "used for psychic communication between organisms", "alien"),
    ("Acid Vat", "used to process prey with corrosive fluids", "alien"),
    ("Spore Chamber", "used to release reproductive spores", "alien"),
    ("Living Gate", "organic structure that opens for hive creatures", "alien"),
    ("Mutation Labyrinth", "used to alter captured organisms", "alien"),
    ("Tentacle Forest", "living terrain used for trapping prey", "alien"),
    ("Brain Throne", "central psychic command node of the hive", "alien"),
]

# Expand with generated entries
templates = {
    "wild": ["Gathering Ground", "Stone Ring", "Hide Shelter", "River Camp", "Bone Cache"],
    "tribal": ["Clan Hut", "Spirit Tent", "Warrior Lodge", "Drum House", "Fishing Village"],
    "medieval": ["Guild Hall", "Stone Keep", "Merchant Row", "Guard Post", "Mason Workshop"],
    "industrial": ["Bus Terminal", "Steel Mill", "Radio Tower", "Parking Garage", "Shipping Port"],
    "ultratech": ["Neural Archive", "Gravity Lab", "Nanite Vault", "Orbital Foundry", "Medical Pod Bay"],
    "spacer": ["Star Beacon", "Jump Gate", "Fuel Depot", "Orbital Habitat", "Survey Vessel"],
    "alien": ["Organic Cathedral", "Larva Chamber", "Hive Spine", "Toxin Pool", "Living Maze"],
}

functions = {
    "wild": "used for survival, shelter, or food gathering",
    "tribal": "used by tribal communities for living, rituals, or defense",
    "medieval": "used by medieval societies for trade, defense, or daily life",
    "industrial": "used for manufacturing, transportation, or urban services",
    "ultratech": "used for advanced science, automation, or futuristic living",
    "spacer": "used for space travel, colonization, or orbital industry",
    "alien": "organic or unnatural structure used by alien lifeforms",
}

for era, names in templates.items():
    for i in range(1, 21):
        for name in names:
            data.append((f"{name} {i}", functions[era], era))

wb = Workbook()
ws = wb.active
ws.title = "Places By Function"

headers = ["name", "function", "tech era"]
ws.append(headers)

for cell in ws[1]:
    cell.font = Font(bold=True)

for row in data:
    ws.append(row)

for col in range(1, 4):
    max_length = 0
    column = get_column_letter(col)
    for cell in ws[column]:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[column].width = min(max_length + 5, 60)

path = "places_by_function.xlsx"
wb.save(path)

print(f"Created spreadsheet with {len(data)} entries.")
print(f"Saved to: {path}")
